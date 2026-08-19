#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""تبدیل فایل اکسل «خرج کلبه» به data/expenses.json

ساختار اکسل: سه جدول کنار هم در یک شیت
  - کارت ملت : ستون A-D ، هدر ردیف ۳ ، داده ردیف ۴ به بعد ، جمع D50
  - سرویس آب : ستون G-J ، هدر ردیف ۳ ، داده ردیف ۴ به بعد ، جمع J16
  - تبلیغات  : ستون M-P ، هدر ردیف ۳ ، داده ردیف ۴ به بعد ، جمع P17

استفاده:
  python scripts/excel_to_json.py <file.xlsx> [-o data/expenses.json]
"""
import sys, json, re, argparse
from pathlib import Path
import openpyxl

# ── برچسب‌های دسته‌بندی ──────────────────────────────────────────────
# شرح‌هایی که باید دقیقاً به دسته معین بروند (بر روی کلمات کلیدی مقدم‌اند)
EXACT_OVERRIDES = {
    'پول شکری منبع آب': 'water',      # دستور کاربر: جزو خرج آب است
}

PERSONAL_EXACT = {'محمد', 'زهرا', 'محمد و زهرا', 'زهرا و محمد', 'زهرا (تو دفتر )', 'زهرا (تو دفتر)'}

def normalize(s):
    return re.sub(r'\s+', ' ', (s or '').strip())

def pick_personal_who(desc):
    """برای ردیف برداشت شخصی: m / z / both"""
    if 'محمد' in desc and 'زهرا' in desc:
        return 'both'
    if 'محمد' in desc:
        return 'm'
    if 'زهرا' in desc:
        return 'z'
    return 'm'

def classify_card(desc):
    """دسته‌بندی ردیف‌های «کارت ملت» با کلمات کلیدی + override صریح"""
    d = normalize(desc)
    if d in EXACT_OVERRIDES:
        return EXACT_OVERRIDES[d]
    if d in PERSONAL_EXACT:
        return 'personal'
    if 'مالیات' in d:
        return 'tax'
    if any(k in d for k in ('شیپور', 'دیوار', 'نردبان', 'آگهی', 'تابلو')):
        return 'ads'
    if any(k in d for k in ('سوادکوهی', 'عزیزی', 'بدهی کلبه', 'شکری محمد رضا')):
        return 'settlement'
    if any(k in d for k in ('مورچه', 'عکس')):
        return 'misc'
    if any(k in d for k in ('مغازه', 'سنگ', 'تخت', 'کولر', 'کلبه', 'دفتر', 'چوب')):
        return 'materials'
    return 'hospitality'

def parse_amount(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    if isinstance(v, str) and v.isdigit():
        return int(v)
    s = re.sub(r'[^\d]', '', str(v))
    return int(s) if s else None

def read_block(ws, row_start, col_row, col_desc, col_date, col_amt):
    """خواندن یک جدول (بلوک) — ردیف‌های با شرح خالی را رد می‌کند"""
    rows = []
    for r in range(row_start, ws.max_row + 1):
        desc = normalize(ws.cell(row=r, column=col_desc).value)
        amt = parse_amount(ws.cell(row=r, column=col_amt).value)
        if not desc or not amt:
            continue
        date = normalize(ws.cell(row=r, column=col_date).value)
        rows.append({'desc': desc, 'date': date, 'amount': amt})
    return rows

def build(ws):
    out = []
    nid = 0

    # کارت ملت (اول — ترتیب دفترچه اصلی)
    card_rows = read_block(ws, 4, 1, 2, 3, 4)     # A/D
    for r in card_rows:
        cat = classify_card(r['desc'])
        row = {'id': nid + 1, 'desc': r['desc'], 'date': r['date'], 'amount': r['amount'],
               'category': cat, 'source': 'card'}
        if cat == 'personal':
            row['who'] = pick_personal_who(r['desc'])
        out.append(row)
        nid += 1

    # آب
    water_rows = read_block(ws, 4, 3, 8, 9, 10)   # G/J هدر ردیف ۳ ولی داده از ۴ — سربرگ بابت/تاریخ/مبلغ را هم در می‌آوریم و حذف می‌کنیم
    water_rows = [r for r in water_rows if r['desc'] not in ('بابت', 'ردیف', 'هزینه', 'مبلغ')]
    for r in water_rows:
        nid += 1
        out.append({'id': nid, 'desc': r['desc'], 'date': r['date'], 'amount': r['amount'],
                    'category': 'water', 'source': 'water'})

    # تبلیغات
    ads_rows = read_block(ws, 4, 13, 14, 15, 16)  # M/P
    for r in ads_rows:
        nid += 1
        out.append({'id': nid, 'desc': r['desc'], 'date': r['date'], 'amount': r['amount'],
                    'category': 'ads', 'source': 'ads'})

    return out

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx', nargs='?', default=r'C:\Users\Ma\AppData\Local\hermes\attachments\خرج کلبه.xlsx')
    ap.add_argument('-o', '--out', default=str(Path(__file__).resolve().parent.parent / 'data' / 'expenses.json'))
    a = ap.parse_args()

    wb = openpyxl.load_workbook(a.xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    data = build(ws)

    out = Path(a.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(data, ensure_ascii=False, indent=1) + '\n', encoding='utf-8')
    print(f'OK {len(data)} rows -> {out}')

if __name__ == '__main__':
    main()
